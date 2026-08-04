# -*- coding: utf-8 -*-
"""
6. Çıktı Modülü (Exporter)
========================
Zenginleştirilmiş ve trend analizi yapılmış tercih verilerini
Excel (.xlsx) ve Markdown (.md) formatlarında kaydeder.
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

class DataExporter:
    @staticmethod
    def export_to_excel(records: List[Dict[str, Any]], output_filepath: str = "yokatlas_sonuclar.xlsx"):
        """Verileri şık bir Excel (.xlsx) tablosu olarak kaydeder."""
        df_rows = []
        for r in records:
            df_rows.append({
                "Veritabanı": r.get("id", ""),
                "Program ID": r.get("program_id", ""),
                "Üniversite & Bölüm Adı": r.get("raw_name", ""),
                "Şehir": r.get("city", ""),
                "Kontenjan / Konum": r.get("location", ""),
                "Geçen Yılki Sıralama": r.get("last_rank", ""),
                "Sıralama Trendi": r.get("trend", ""),
                "Tahmini Skor": r.get("tahmin", ""),
                "Güven Aralığı (%10)": r.get("range_str", ""),
                "Kişisel Puanım (1-10)": r.get("rating", ""),
                "Notlar / Artılar - Eksiler": r.get("notes", "")
            })

        df = pd.DataFrame(df_rows)
        df.to_excel(output_filepath, index=False, engine="openpyxl")

        # OpenPyXL Stil Giydirme
        wb = openpyxl.load_workbook(output_filepath)
        ws = wb.active

        # Stiller
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10)
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        # Başlık Satırı
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Veri Satırları
        for row_idx in range(2, len(records) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = thin_border

                # Hiza ayarları
                if col_idx in [1, 2, 4, 6, 8, 9, 10]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

                # Öne çıkan sütun vurguları
                if col_idx == 8: # Tahmini Skor
                    cell.font = bold_font
                    cell.fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

        # Otomatik Kolon Genişliği
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output_filepath)
        print(f"✅ Excel raporu başarıyla kaydedildi: {os.path.abspath(output_filepath)}")

    @staticmethod
    def export_to_markdown(records: List[Dict[str, Any]], output_filepath: str):
        """Verileri Markdown (.md) tercih tablosu olarak kaydeder."""
        title = "# YKS Zenginleştirilmiş Tercih & Tahmin Analiz Tablosu\n\n"
        header = "| Veritabanı | Üniversite & Bölüm Adı | Şehir | Kontenjan / Konum | Geçen Yılki Sıralama | Kişisel Puanım (1-10) | Notlar / Artılar - Eksiler | Sıralama Trendi | Tahmini Skor |\n"
        divider = "| :---: | :--- | :--- | :--- | :---: | :---: | :--- | :---: | :---: |\n"

        lines = [title, header, divider]

        for r in records:
            db_id = r.get("id", "")
            raw_name = r.get("raw_name", "")
            city = r.get("city", "")
            location = r.get("location", "") or city
            last_rank = r.get("last_rank", "-")
            rating = r.get("rating", "-")
            notes = r.get("notes", "-")
            trend = r.get("trend", "Yatay ➡️")
            tahmin = r.get("tahmin", "-")

            line = f"| **{db_id}** | {raw_name} | {city} | {location} | {last_rank} | {rating} | {notes} | {trend} | {tahmin} |\n"
            lines.append(line)

        # Check if there are any detailed AI analyses (ai_eval) to append
        has_ai_evals = any(r.get("ai_eval") for r in records)
        if has_ai_evals:
            lines.append("\n---\n\n## 🤖 Detaylı Yapay Zeka Analiz ve Değerlendirme Raporları\n\n")
            for r in records:
                ai_eval_html = r.get("ai_eval")
                if ai_eval_html:
                    db_id = r.get("id", "")
                    raw_name = r.get("raw_name", "")
                    lines.append(f"### {db_id}. {raw_name}\n\n")
                    
                    # Convert simple HTML back to markdown for clean reading in .md files
                    md_text = ai_eval_html
                    md_text = md_text.replace("<p>", "").replace("</p>", "\n\n")
                    md_text = md_text.replace("<strong>", "**").replace("</strong>", "**")
                    md_text = md_text.replace("<ul>", "").replace("</ul>", "\n")
                    md_text = md_text.replace("<li>", "- ").replace("</li>", "\n")
                    
                    # Clean up double/triple newlines
                    while "\n\n\n" in md_text:
                        md_text = md_text.replace("\n\n\n", "\n\n")
                        
                    lines.append(md_text.strip() + "\n\n")

        with open(output_filepath, "w", encoding="utf-8") as f:
            f.writelines(lines)

        print(f"✅ Markdown raporu kaydedildi: {os.path.abspath(output_filepath)}")
